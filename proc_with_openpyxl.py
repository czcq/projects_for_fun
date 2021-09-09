from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook(filename='/home/yang/桌面/test.xlsx')

# 在这里指定需要处理哪两列
first_column_index = 'C'
second_column_index = 'F'

tmp_dict = {}

i = 0
for item in wb.active[first_column_index]:
    key = item.value
    list_element = {}
    list_element['coordinate'] = wb.active[second_column_index][i].coordinate
    list_element['value'] = wb.active[second_column_index][i].value
    if key in tmp_dict:
        tmp_dict[key].append(list_element)
    else:
        value = [list_element]
        tmp_dict[key] = value
    i = i + 1

# merge
for key in tmp_dict:
    length = len(tmp_dict[key])
    # the cell need merge
    if length > 1:
        merge_cells_value_list = []
        for list_element in tmp_dict[key]:
            merge_cells_value_list.append(str(list_element['value']))
        merge_cells_value = '\n'.join(merge_cells_value_list)
        fist_cell_coordinate = tmp_dict[key][0]['coordinate']
        last_cell_coordinate = tmp_dict[key][length - 1]['coordinate']
        wb.active.merge_cells(fist_cell_coordinate + ":" + last_cell_coordinate)
        print(key)
        first_cell = wb.active[fist_cell_coordinate]
        first_cell.alignment = Alignment(wrap_text=True)
        first_cell.value = merge_cells_value


wb.save(filename="output.xlsx")
